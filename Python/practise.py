# #################making a game to guess secret number##############################################
# secret_number = 10
# guess_count = 0
# guess_limit = 3
# while guess_count < guess_limit :
#     guess = int(input("Enter your guess")) 
#     guess_count+= 1
#     if guess == secret_number:
#         print("You guessed it right")
#         break
#     else:
#         print("Wrong guess")
# else:
#     print("sorry! you failed")
# ############################trying to make a simple game##################not working################
## whats the problem in writing this code with if and not while as done below
# command = input("What do you wanna do") 
# start =""
# stop = ""

# if command == help:
#     print("'start' - to start the car"
#           "'stop' - to stop the car "
#           "'quit' - to quit the game")
#     if command == start:
#         print("the car has started. Initialising engine")
#     elif command == stop :
#         print("the car is stopped. Initialising breaks")
#     else:
#         print("game stopped")
        
# else :
#     print("write a valid command")

####################made a simple game##########################################

# command = ""
# started =False

# while True:
#     command = input("write your command ->").lower()
    
#     if command == "start":        
#         if started :
#             print("car is already started")
#         else:
#             started= True
#             print("Initialising engine")
            
            
#     elif command == "stop" :
#         if not started:
#             print("Car is already stopped")
#         else:
#             started = False
#             print("Initialising breaks")  
               
#     elif command == "help":
#         print(""" 
#             'start' - to start the car
#             'stop' - to stop the car 
#             'quit' - to quit the game 
#             """)
#     elif command == "quit":
#        break
   
#     else:
#         print("Invalid command")
    
#  # https://youtu.be/_uQrJ0TkZlc?t=5809  (should else be in 'while block' or 'if block')
######################making shape 'F' using python########################################

# number = [5,2,5,2,2]
# for i in number:     # or just write : print ("x"* i)
#     output = ""
#     for j in range(i):
#         output += 'x'
#     print(output)

##########################finding out the largest number####################################


# num = [1,2,5,6,90,63,46]
# max = num[0]
# for number in num:
#     if number > max:
#         max=number
# print(f'largest number is {max}')
################ the below Question is created and answered by me ##########################
# names = ['aidskhbc','rjhvb','hbsdkb','khbvd']
# max = len(names[0])
# for letters in names:
#     if len(letters) > max:
#         max = letters
# print(max)

#########################removing duplicate numbers#####################################
# numbers = [1,53,64,24,79,64,24,24]
# for i in numbers:
#     if numbers.index(i) > 1:
#         numbers.remove(i)
# print(numbers)

# numbers = [1,53,64,24,79,64,24,24]
# uniques = []
# for i in numbers:
#     if i not in uniques:
#         uniques.append(i)
# print(uniques)
# #############learn the logic of this code

###################changing phone numbers to words############################################
# phone = input("Phone : ") ## why we dont take value as int()
# naming ={
#     "1" : "one",
#     "2": "two",
#     "3" : "three",
#     "4" : "four",
#     "5": "five",
#     "6" : "six",
#     "7" : "seven",
#     "8": "eight",
#     "9" : "nine",    
# }
# output = ""
# for ch in phone:
#     output += naming.get(ch, "x") + " "
# print(output)


########################Emoji converter#######################################
# message = input(">")
# words = message.split(" ")
# emoji ={
#     ":)" : "😀",
#     ":(" : "😢",
#     ":/" : "😕"
# }
# output = ""
# for word in words:    
#     output += emoji.get( word , word) + " "
# print(output)

#########################emoji v2 using function######################################
# def emoji(message) :
#     words = message.split(" ")
#     emoji ={
#         ":)" : "😀",
#         ":(" : "😢",
#         ":/" : "😕"
#     }
#     output = ""
#     for word in words:    
#         output += emoji.get( word , word) + " "
#         return output


# message = input(">")
# print(emoji(message))

###############################################################

# class Person:
#     def __init__(self , name):
#         self.name = name
    
#     def talk(self):
#         print(f"Hi, I am {self .name}")
    
# sam = Person("Sam testing")
# sam.talk()

#####################Max number v2 using modules##########################################
# from conversion import find_max

# numbers = [10,3,56,43]

# max = find_max(numbers)
# print(max)

########################Genrating random dice values#######################################
# import random


# class Dice:
#     def roll(self):
#        first= random.randint(1,6)
#        second= random.randint(1,6)
#        return first, second
   

# dice = Dice()
# print( dice.roll())

###########################Automating excel file####################################

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1,1)
    

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
        
    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4
            )    

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")


    wb.save(filename)




            